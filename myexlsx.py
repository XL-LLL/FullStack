
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell,MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill
from openpyxl.styles import Side
work_obj = load_workbook('E:\Code\FullStack\icon\myxlsx.xlsx')

sheetname = work_obj.sheetnames
print(sheetname)
#h获取sheet对象
work_sheet = work_obj.worksheets
print(work_sheet[0])
work_sheet0 = work_sheet[0]
#u获取单元格
cell_obj =work_sheet0.cell(1,1)
print(cell_obj.value)
print(work_sheet0["A1"].value)
#读取一行
row_list = work_sheet0[1]
for row in row_list:
    print(row.value)

#获取所有行
for row in work_sheet0.rows:
    list = []
    for cell in row:
        list.append(cell.value)
    print(list)
#获取某一列
for row in work_sheet0.rows:
    list = []
    cell = row[0]
    print(cell.value)

#合并的单元格 被合并的单元格是mergedcell 是没有值的
for row in work_sheet0.rows:
    list = []
    for cell in row:
        if type(cell) == MergedCell:
            list.append("--")
        elif type(cell) == Cell:
            list.append(cell.value)
    print(list)
#将被合并的单元格真实内容显示  比如找c3
for row in work_sheet0.merged_cells:
    print(row )#返回的是范围
    if "C3" in row:
        print(row.start_cell.value)#获取头部的函数

#获取cell的坐标
print(cell_obj.coordinate)
#获取一共有多少行多少列
print(work_sheet0.max_row)
print(work_sheet0.max_column)


#从第n行开始读取
for row in work_sheet0.iter_rows(min_row=2, max_row=3):
    list = []
    for cell in row:
        list.append(cell.value)
    print(list)


#写入内容
print(cell_obj.value)
cell_obj.value = "修改"
work_obj.save('E:\Code\FullStack\icon\myxlsx.xlsx')

#新建就是写入一个空的保存
from openpyxl.workbook import Workbook
workbook = Workbook()
sheet_obj = workbook.worksheets[0]
cell_objnwe = sheet_obj.cell(1,1)
cell_objnwe.value = "hahah"
cell_objnwe.alignment = Alignment(horizontal='center', vertical='center')
cell_objnwe.border = Border(left=Side(border_style='medium', color='0000FF'))
cell_objnwe.font = Font(color='0000FF',size=40,name="宋体")
cell_objnwe.fill = PatternFill("solid",fgColor="00FFFF")

sheet_obj.title = "lalal"

sheet_obj.row_dimensions[2].height = 30

sheet_obj.merge_cells('A2:D3')


sheet_obj["A4"] = 10
sheet_obj["A5"] = 20
sheet_obj["A6"] ="=A4*A5"
sheet_obj["A7"] ="=SUM(A4,A5,A6)"

workbook.create_sheet("mamam")
workbook.copy_worksheet(sheet_obj)


workbook.save('E:\Code\FullStack\icon\mynew.xlsx')